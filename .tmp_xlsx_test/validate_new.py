#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证新 XlsxWriter 结构：目录树 + 标准 zip（模拟系统 zip）+ STORE 回退"""
import io, os, struct, zipfile, shutil
from PIL import Image

BASE = '/workspace/.tmp_xlsx_test'
TREE = os.path.join(BASE, 'pkg')

def build_tree():
    shutil.rmtree(TREE, ignore_errors=True)
    def w(rel, content):
        p = os.path.join(TREE, rel)
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, 'wb') as f:
            f.write(content if isinstance(content, bytes) else content.encode('utf-8'))
    w('[Content_Types].xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Default Extension="jpeg" ContentType="image/jpeg"/>'
        '<Default Extension="jpg" ContentType="image/jpeg"/>'
        '<Default Extension="png" ContentType="image/png"/>'
        '<Default Extension="gif" ContentType="image/gif"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>')
    w('_rels/.rels', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>')
    w('docProps/app.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties">'
        '<Application>WatermarkCamera</Application></Properties>')
    w('docProps/core.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        '<dc:creator>WatermarkCamera</dc:creator>'
        '<dcterms:created xsi:type="dcterms:W3CDTF">2026-08-30T08:00:00Z</dcterms:created>'
        '<dcterms:modified xsi:type="dcterms:W3CDTF">2026-08-30T08:00:00Z</dcterms:modified>'
        '</cp:coreProperties>')
    w('xl/workbook.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<bookViews><workbookView activeTab="0"/></bookViews>'
        '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>')
    w('xl/_rels/workbook.xml.rels', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '</Relationships>')
    w('xl/styles.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<numFmts count="1"><numFmt numFmtId="164" formatCode="@"/></numFmts>'
        '<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font>'
        '<font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font></fonts>'
        '<fills count="3"><fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FF4472C4"/><bgColor indexed="64"/></patternFill></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="3">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="0"/>'
        '<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1" applyAlignment="1"><alignment vertical="center" wrapText="1"/></xf>'
        '<xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>'
        '</cellXfs><cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>')
    w('xl/worksheets/sheet1.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<cols><col min="1" max="1" width="32" customWidth="1"/><col min="2" max="5" width="13" customWidth="1"/></cols>'
        '<sheetData>'
        '<row r="1" ht="21" customHeight="1">'
        '<c r="A1" s="2" t="inlineStr"><is><t>FOTO</t></is></c>'
        '<c r="B1" s="2" t="inlineStr"><is><t>CODIGO</t></is></c>'
        '<c r="C1" s="2" t="inlineStr"><is><t>DETALLADOS</t></is></c>'
        '<c r="D1" s="2" t="inlineStr"><is><t>描述</t></is></c>'
        '<c r="E1" s="2" t="inlineStr"><is><t>PRECIO</t></is></c>'
        '</row>'
        '<row r="2" ht="135" customHeight="1">'
        '<c r="A2" s="1"/><c r="B2" s="1" t="inlineStr"><is><t xml:space="preserve">M001</t></is></c>'
        '<c r="C2" s="1" t="inlineStr"><is><t xml:space="preserve">algún texto español</t></is></c>'
        '<c r="D2" s="1" t="inlineStr"><is><t xml:space="preserve">中文描述</t></is></c>'
        '<c r="E2" s="1" t="inlineStr"><is><t xml:space="preserve">12.50</t></is></c>'
        '</row>'
        '</sheetData><drawing r:id="rId3"/></worksheet>')
    w('xl/worksheets/_rels/sheet1.xml.rels', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
        '</Relationships>')
    w('xl/drawings/drawing1.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<xdr:twoCellAnchor>'
        '<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
        '<xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>2</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="Image 1"/><xdr:cNvPicPr/></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip r:embed="rId1"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
        '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="2095500" cy="1571625"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
        '</xdr:pic><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>')
    w('xl/drawings/_rels/drawing1.xml.rels', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.jpeg"/>'
        '</Relationships>')
    im = Image.new('RGB', (800, 600), (60, 120, 200))
    b = io.BytesIO(); im.save(b, 'JPEG', quality=85)
    w('xl/media/image1.jpeg', b.getvalue())

def zip_deflate(src, dest):  # 模拟系统（Apple）zip：DEFLATE + 真实时间戳
    with zipfile.ZipFile(dest, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, dirs, files in os.walk(src):
            for f in sorted(files):
                full = os.path.join(root, f)
                z.write(full, os.path.relpath(full, src))

CRC_TABLE = []
for n in range(256):
    c = n
    for _ in range(8):
        c = (0xEDB88320 ^ (c >> 1)) if (c & 1) else (c >> 1)
    CRC_TABLE.append(c)

def crc32(data):
    crc = 0xFFFFFFFF
    for b in data:
        crc = CRC_TABLE[(crc ^ b) & 0xFF] ^ (crc >> 8)
    return crc ^ 0xFFFFFFFF

def zip_store_fixed(src, dest):  # 模拟回退：STORE + 合法 DOS 日期(0x21)/flags 0/vmb 20
    buf = io.BytesIO(); central = io.BytesIO(); offset = 0; count = 0
    files = []
    for root, dirs, fs in os.walk(src):
        for f in fs:
            files.append(os.path.join(root, f))
    files.sort()
    for full in files:
        rel = os.path.relpath(full, src).replace(os.sep, '/')
        with open(full, 'rb') as fh:
            data = fh.read()
        nb = rel.encode(); crc = crc32(data)
        buf.write(struct.pack('<IHHHHHIIIHH', 0x04034b50, 20, 0, 0, 0, 0x21,
                              crc, len(data), len(data), len(nb), 0) + nb)
        buf.write(data)
        central.write(struct.pack('<IHHHHHHIIIHHHHHII', 0x02014b50, 0x0014, 20, 0, 0, 0, 0x21,
                                  crc, len(data), len(data), len(nb), 0, 0, 0, 0, 0, offset) + nb)
        offset += 30 + len(nb) + len(data); count += 1
    buf.write(central.getvalue())
    cd = central.tell()
    buf.write(struct.pack('<IHHHHIIH', 0x06054b50, 0, 0, count, count, cd, offset, 0))
    with open(dest, 'wb') as f:
        f.write(buf.getvalue())

if __name__ == '__main__':
    build_tree()
    zip_deflate(TREE, os.path.join(BASE, 'new_deflate.xlsx'))
    zip_store_fixed(TREE, os.path.join(BASE, 'new_store.xlsx'))
    print('generated')

    from openpyxl import load_workbook
    for name in ['new_deflate.xlsx', 'new_store.xlsx']:
        p = os.path.join(BASE, name)
        wb = load_workbook(p)
        ws = wb.active
        assert ws['B2'].value == 'M001' and ws['D2'].value == '中文描述'
        assert len(ws._images) == 1
        print('%s: openpyxl OK (%d parts)' % (name, len(zipfile.ZipFile(p).namelist())))
